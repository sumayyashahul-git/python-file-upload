# excel_utils.py
import openpyxl

def copy_excel_data():
    source_file = r'C:\Users\user\Desktop\Leavebalance Mismatch.xlsx'
    wb_source = openpyxl.load_workbook(source_file)
    sheet_source = wb_source.active

    wb_output = openpyxl.Workbook()
    sheet_output = wb_output.active
    sheet_output.title = "FilteredData"

    for row in sheet_source.iter_rows(values_only=True):
        sheet_output.append(row)

    output_file = r'C:\Users\user\Desktop\Output.xlsx'
    wb_output.save(output_file)

    wb_source.close()
    wb_output.close()

    print(f"Done! Data written to: {output_file}")
